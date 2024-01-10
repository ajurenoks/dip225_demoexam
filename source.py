from openpyxl import Workbook, load_workbook 
wb=load_workbook('sagatave.xlsx')
ws=wb.active
max_row=ws.max_row
s=[]
for row in range(2,max_row+1):
    a=(ws['a' + str(row)].value)
    s.append(a)
print((s))
